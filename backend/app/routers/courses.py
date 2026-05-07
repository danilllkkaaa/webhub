from datetime import datetime
import io
import secrets
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, distinct, delete
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from slugify import slugify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.user import User
from app.models.course import (
    Course, CourseModule, CourseLesson, CourseStudent, 
    CourseLessonProgress, CourseStatus, CourseStudentStatus,
    Quiz, QuizQuestion, QuizOption, QuizAttempt, QuizQuestionType
)
from app.schemas.course import (
    CourseCreate, CourseUpdate, CourseOut, CourseStructureOut,
    ModuleCreate, ModuleUpdate, ModuleOut,
    LessonCreate, LessonUpdate, LessonOut,
    CourseJoinRequest, CourseStudentOut, CourseLearnOut,
    QuizCreate, QuizUpdate, QuizOut, QuizFullOut, QuizAttemptCreate, QuizAttemptOut
)
from app.core.dependencies import get_current_user
from app.core.security import generate_viewer_token, hash_password, verify_password
from app.services.access import (
    accessible_project_ids,
    get_course_for_user,
    get_course_lesson_for_user,
    get_course_module_for_user,
    optional_project_id_for_user,
    require_organization_id,
)

router = APIRouter(tags=["courses"])


def _quiz_load_options():
    return selectinload(CourseModule.quiz).selectinload(Quiz.questions).selectinload(QuizQuestion.options)


def _quiz_options():
    return selectinload(Quiz.questions).selectinload(QuizQuestion.options)


async def _unique_slug(title: str, db: AsyncSession) -> str:
    base = slugify(title) or "course"
    slug = base
    i = 1
    while True:
        result = await db.execute(select(Course).where(Course.slug == slug))
        if not result.scalar_one_or_none():
            return slug
        slug = f"{base}-{i}"
        i += 1


async def _unique_public_id(db: AsyncSession) -> str:
    while True:
        public_id = str(secrets.randbelow(900000) + 100000)
        result = await db.execute(select(Course.id).where(Course.public_id == public_id))
        if result.scalar_one_or_none() is None:
            return public_id


async def _course_for_user_by_ref(course_ref: str, user: User, db: AsyncSession, permission: str = "course.view") -> Course:
    if course_ref.isdigit() and len(course_ref) != 6:
        return await get_course_for_user(int(course_ref), user, db, permission)

    organization_id = require_organization_id(user)
    result = await db.execute(
        select(Course).where(
            Course.public_id == course_ref,
            Course.organization_id == organization_id,
        )
    )
    course = result.scalar_one_or_none()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    if course.id is None:
        raise HTTPException(status_code=404, detail="Course not found")
    await get_course_for_user(course.id, user, db, permission)
    return course


async def _course_out(course: Course, db: AsyncSession) -> CourseOut:
    module_count = (await db.execute(select(func.count(CourseModule.id)).where(CourseModule.course_id == course.id))).scalar() or 0
    lesson_count = (
        await db.execute(
            select(func.count(CourseLesson.id))
            .join(CourseModule, CourseLesson.module_id == CourseModule.id)
            .where(CourseModule.course_id == course.id)
        )
    ).scalar() or 0
    student_count = (await db.execute(select(func.count(CourseStudent.id)).where(CourseStudent.course_id == course.id))).scalar() or 0
    return CourseOut(
        id=course.id,
        public_id=course.public_id or f"{course.id:06d}",
        project_id=course.project_id,
        slug=course.slug,
        invite_token=course.invite_token,
        title=course.title,
        description=course.description,
        cover_url=course.cover_url,
        status=course.status,
        sequential_access=course.sequential_access,
        module_count=module_count,
        lesson_count=lesson_count,
        student_count=student_count,
        created_at=course.created_at,
        updated_at=course.updated_at,
    )


async def _student_out(student: CourseStudent, db: AsyncSession) -> CourseStudentOut:
    total = (
        await db.execute(
            select(func.count(CourseLesson.id))
            .join(CourseModule, CourseLesson.module_id == CourseModule.id)
            .where(CourseModule.course_id == student.course_id)
            .where(CourseModule.is_published.is_(True))
            .where(CourseLesson.is_published.is_(True))
        )
    ).scalar() or 0
    completed = (
        await db.execute(
            select(func.count(distinct(CourseLessonProgress.lesson_id))).where(CourseLessonProgress.student_id == student.id)
        )
    ).scalar() or 0
    percent = round(completed / total * 100, 1) if total else 0
    return CourseStudentOut(
        id=student.id,
        course_id=student.course_id,
        token=student.token,
        name=student.name,
        phone=student.phone,
        email=student.email,
        telegram=student.telegram,
        status=student.status,
        progress_percent=percent,
        completed_lessons=completed,
        total_lessons=total,
        created_at=student.created_at,
        last_seen_at=student.last_seen_at,
    )


async def _published_lessons_for_course(course_id: int, db: AsyncSession) -> list[CourseLesson]:
    result = await db.execute(
        select(CourseLesson)
        .join(CourseModule, CourseLesson.module_id == CourseModule.id)
        .where(CourseModule.course_id == course_id)
        .where(CourseModule.is_published.is_(True))
        .where(CourseLesson.is_published.is_(True))
        .order_by(CourseModule.position, CourseLesson.position, CourseLesson.id)
    )
    return list(result.scalars().all())


async def _ensure_student_can_learn(student: CourseStudent) -> None:
    if student.status != CourseStudentStatus.approved:
        raise HTTPException(status_code=403, detail="Course access has not been approved yet")


async def _completed_lesson_ids(student_id: int, db: AsyncSession) -> set[int]:
    completed = await db.execute(
        select(CourseLessonProgress.lesson_id).where(CourseLessonProgress.student_id == student_id)
    )
    return set(completed.scalars().all())


async def _ensure_lesson_available(
    course: Course,
    student: CourseStudent,
    lesson: CourseLesson,
    db: AsyncSession,
) -> None:
    lessons = await _published_lessons_for_course(course.id, db)
    lesson_ids = [item.id for item in lessons]
    if lesson.id not in lesson_ids:
        raise HTTPException(status_code=404, detail="Lesson not found")
    if not course.sequential_access:
        return
    lesson_index = lesson_ids.index(lesson.id)
    previous_ids = set(lesson_ids[:lesson_index])
    completed_ids = await _completed_lesson_ids(student.id, db)
    if not previous_ids.issubset(completed_ids):
        raise HTTPException(status_code=403, detail="Предыдущие уроки еще не завершены.")


async def _ensure_quiz_available(
    course: Course,
    student: CourseStudent,
    quiz: Quiz,
    db: AsyncSession,
) -> None:
    if not course.sequential_access:
        return
    module_lessons = await db.execute(
        select(CourseLesson.id)
        .where(CourseLesson.module_id == quiz.module_id)
        .where(CourseLesson.is_published.is_(True))
        .order_by(CourseLesson.position, CourseLesson.id)
    )
    required_lesson_ids = set(module_lessons.scalars().all())
    completed_ids = await _completed_lesson_ids(student.id, db)
    if not required_lesson_ids.issubset(completed_ids):
        raise HTTPException(status_code=403, detail="Перед тестом нужно завершить уроки модуля.")


def _validate_quiz_payload(data: QuizCreate) -> None:
    if not data.title.strip():
        raise HTTPException(status_code=422, detail="Quiz title is required")
    if not data.questions:
        raise HTTPException(status_code=422, detail="Quiz must have at least one question")

    for index, question in enumerate(data.questions, start=1):
        if not question.text.strip():
            raise HTTPException(status_code=422, detail=f"Question {index} text is required")
        if len(question.options) < 2:
            raise HTTPException(status_code=422, detail=f"Question {index} must have at least two options")
        if any(not option.text.strip() for option in question.options):
            raise HTTPException(status_code=422, detail=f"Question {index} has an empty option")

        correct_count = sum(1 for option in question.options if option.is_correct)
        if question.question_type == QuizQuestionType.single and correct_count != 1:
            raise HTTPException(status_code=422, detail=f"Question {index} must have exactly one correct option")
        if question.question_type == QuizQuestionType.multiple and correct_count < 1:
            raise HTTPException(status_code=422, detail=f"Question {index} must have at least one correct option")


# --- Courses ---

@router.get("/courses/", response_model=list[CourseOut])
async def list_courses(
    project_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    organization_id = require_organization_id(user)
    query = select(Course).where(Course.organization_id == organization_id).order_by(Course.created_at.desc())
    allowed_project_ids = await accessible_project_ids(user, db)
    if allowed_project_ids is not None:
        query = query.where(Course.project_id.in_(allowed_project_ids))
    if project_id is not None:
        await optional_project_id_for_user(project_id, user, db, "course.view")
        query = query.where(Course.project_id == project_id)
    result = await db.execute(query)
    return [await _course_out(course, db) for course in result.scalars().all()]


@router.post("/courses/", response_model=CourseOut, status_code=201)
async def create_course(
    data: CourseCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    organization_id = require_organization_id(user)
    project_id = await optional_project_id_for_user(data.project_id, user, db, "course.create")
    course = Course(
        organization_id=organization_id,
        project_id=project_id,
        public_id=await _unique_public_id(db),
        slug=await _unique_slug(data.title, db),
        invite_token=generate_viewer_token(),
        title=data.title.strip(),
        description=data.description,
        cover_url=data.cover_url,
        status=data.status,
        sequential_access=data.sequential_access,
    )
    db.add(course)
    await db.commit()
    await db.refresh(course)
    return await _course_out(course, db)


@router.get("/courses/{course_ref}", response_model=CourseOut)
async def get_course(course_ref: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    return await _course_out(await _course_for_user_by_ref(course_ref, user, db), db)


@router.patch("/courses/{course_ref}", response_model=CourseOut)
async def update_course(
    course_ref: str,
    data: CourseUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    course = await _course_for_user_by_ref(course_ref, user, db, "course.update")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(course, field, value)
    await db.commit()
    await db.refresh(course)
    return await _course_out(course, db)


@router.delete("/courses/{course_ref}", status_code=204)
async def delete_course(course_ref: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    course = await _course_for_user_by_ref(course_ref, user, db, "course.delete")
    await db.delete(course)
    await db.commit()


@router.get("/courses/{course_ref}/structure", response_model=CourseStructureOut)
async def course_structure(course_ref: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    course = await _course_for_user_by_ref(course_ref, user, db)
    modules_result = await db.execute(
        select(CourseModule)
        .where(CourseModule.course_id == course.id)
        .options(_quiz_load_options())
        .order_by(CourseModule.position, CourseModule.id)
    )
    modules = modules_result.scalars().all()
    
    lessons_result = await db.execute(
            select(CourseLesson)
            .join(CourseModule, CourseLesson.module_id == CourseModule.id)
            .where(CourseModule.course_id == course.id)
        .order_by(CourseModule.position, CourseLesson.position, CourseLesson.id)
    )
    lessons = lessons_result.scalars().all()
    
    return CourseStructureOut(course=await _course_out(course, db), modules=modules, lessons=lessons)


# --- Modules ---

@router.post("/courses/{course_ref}/modules", response_model=ModuleOut, status_code=201)
async def create_module(course_ref: str, data: ModuleCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    course = await _course_for_user_by_ref(course_ref, user, db, "course.update")
    module = CourseModule(course_id=course.id, **data.model_dump())
    db.add(module)
    await db.commit()
    await db.refresh(module)
    
    # Re-fetch to ensure relationship attributes don't trigger lazy load
    res = await db.execute(
        select(CourseModule).where(CourseModule.id == module.id).options(_quiz_load_options())
    )
    return res.scalar_one()


@router.patch("/courses/modules/{module_id}", response_model=ModuleOut)
async def update_module(module_id: int, data: ModuleUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    await get_course_module_for_user(module_id, user, db, "course.update")
    result = await db.execute(
        select(CourseModule)
        .where(CourseModule.id == module_id)
        .options(_quiz_load_options())
    )
    module = result.scalar_one_or_none()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(module, field, value)
    
    await db.commit()
    
    # Re-fetch to ensure relationship attributes don't trigger lazy load
    res = await db.execute(
        select(CourseModule).where(CourseModule.id == module_id).options(_quiz_load_options())
    )
    return res.scalar_one()


@router.delete("/courses/modules/{module_id}", status_code=204)
async def delete_module(module_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    module = await get_course_module_for_user(module_id, user, db, "course.update")
    await db.delete(module)
    await db.commit()


# --- Lessons ---

@router.post("/courses/modules/{module_id}/lessons", response_model=LessonOut, status_code=201)
async def create_lesson(module_id: int, data: LessonCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    module = await get_course_module_for_user(module_id, user, db, "course.update")
    payload = data.model_dump()
    lesson = CourseLesson(module_id=module_id, **payload)
    db.add(lesson)
    await db.commit()
    await db.refresh(lesson)
    return lesson


@router.patch("/courses/lessons/{lesson_id}", response_model=LessonOut)
async def update_lesson(lesson_id: int, data: LessonUpdate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    lesson = await get_course_lesson_for_user(lesson_id, user, db, "course.update")
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(lesson, field, value)
    await db.commit()
    await db.refresh(lesson)
    return lesson


@router.delete("/courses/lessons/{lesson_id}", status_code=204)
async def delete_lesson(lesson_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    lesson = await get_course_lesson_for_user(lesson_id, user, db, "course.update")
    await db.delete(lesson)
    await db.commit()


# --- Quizzes (Admin) ---

@router.post("/courses/modules/{module_id}/quiz", response_model=QuizFullOut, status_code=201)
async def create_quiz(module_id: int, data: QuizCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    await get_course_module_for_user(module_id, user, db, "course.update")
    _validate_quiz_payload(data)
    # Check if quiz exists
    existing = (await db.execute(select(Quiz).where(Quiz.module_id == module_id))).scalar_one_or_none()
    if existing:
        raise HTTPException(status_code=400, detail="Quiz already exists for this module")
    
    quiz = Quiz(module_id=module_id, title=data.title, description=data.description, passing_score=data.passing_score)
    db.add(quiz)
    await db.flush()

    for q_data in data.questions:
        question = QuizQuestion(quiz_id=quiz.id, text=q_data.text, question_type=q_data.question_type, position=q_data.position)
        db.add(question)
        await db.flush()
        for o_data in q_data.options:
            db.add(QuizOption(question_id=question.id, text=o_data.text, is_correct=o_data.is_correct, position=o_data.position))
    
    await db.commit()
    await db.refresh(quiz)
    res = await db.execute(select(Quiz).where(Quiz.id == quiz.id).options(_quiz_options()))
    return res.scalar_one()


@router.put("/courses/modules/{module_id}/quiz", response_model=QuizFullOut)
async def update_quiz(module_id: int, data: QuizCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    await get_course_module_for_user(module_id, user, db, "course.update")
    _validate_quiz_payload(data)
    result = await db.execute(select(Quiz).where(Quiz.module_id == module_id))
    quiz = result.scalar_one_or_none()
    
    if not quiz:
        # If it doesn't exist, create it (acts like an upsert)
        return await create_quiz(module_id, data, db, user)

    # Update base fields
    quiz.title = data.title
    quiz.description = data.description
    quiz.passing_score = data.passing_score
    
    # Delete old questions (cascade will delete options)
    await db.execute(delete(QuizQuestion).where(QuizQuestion.quiz_id == quiz.id))
    await db.flush()

    # Recreate questions and options
    for q_data in data.questions:
        question = QuizQuestion(quiz_id=quiz.id, text=q_data.text, question_type=q_data.question_type, position=q_data.position)
        db.add(question)
        await db.flush()
        for o_data in q_data.options:
            db.add(QuizOption(question_id=question.id, text=o_data.text, is_correct=o_data.is_correct, position=o_data.position))
    
    await db.commit()
    await db.refresh(quiz)
    
    res = await db.execute(select(Quiz).where(Quiz.id == quiz.id).options(_quiz_options()))
    return res.scalar_one()

@router.get("/courses/modules/{module_id}/quiz", response_model=QuizFullOut)
async def get_quiz(module_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    await get_course_module_for_user(module_id, user, db, "course.view")
    quiz = (await db.execute(select(Quiz).where(Quiz.module_id == module_id).options(_quiz_options()))).scalar_one_or_none()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    return quiz


@router.delete("/courses/modules/{module_id}/quiz", status_code=204)
async def delete_quiz(module_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    await get_course_module_for_user(module_id, user, db, "course.update")
    quiz = (await db.execute(select(Quiz).where(Quiz.module_id == module_id))).scalar_one_or_none()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    await db.delete(quiz)
    await db.commit()


# --- Students ---

@router.get("/courses/{course_ref}/students", response_model=list[CourseStudentOut])
async def list_students(course_ref: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    course = await _course_for_user_by_ref(course_ref, user, db, "course.students")
    result = await db.execute(select(CourseStudent).where(CourseStudent.course_id == course.id).order_by(CourseStudent.created_at.desc()))
    return [await _student_out(student, db) for student in result.scalars().all()]


@router.get("/courses/{course_ref}/students/export.xlsx")
async def export_students_xlsx(course_ref: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    course = await _course_for_user_by_ref(course_ref, user, db, "course.export")
    result = await db.execute(select(CourseStudent).where(CourseStudent.course_id == course.id).order_by(CourseStudent.created_at))
    students = [await _student_out(student, db) for student in result.scalars().all()]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ученики"
    headers = ["ФИО", "Телефон", "Email", "Telegram", "Прогресс", "Пройдено уроков", "Всего уроков", "Дата входа", "Последний вход"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="EEF2FF")
    for s in students:
        sheet.append([
            s.name, s.phone, s.email, s.telegram, f"{s.progress_percent}%",
            s.completed_lessons, s.total_lessons,
            s.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            s.last_seen_at.strftime("%Y-%m-%d %H:%M:%S") if s.last_seen_at else "",
        ])
    for index, width in enumerate([32, 20, 32, 22, 14, 18, 14, 22, 22], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="course_students_{course.public_id or course.id}.xlsx"'},
    )


# --- Public / Learn ---

@router.get("/course/invite/{invite_token}", response_model=CourseOut)
async def public_course_invite(invite_token: str, db: AsyncSession = Depends(get_db)):
    course = (await db.execute(select(Course).where(Course.invite_token == invite_token))).scalar_one_or_none()
    if not course or course.status == CourseStatus.archived:
        raise HTTPException(status_code=404, detail="Course not found")
    return await _course_out(course, db)


@router.post("/course/invite/{invite_token}/join", response_model=CourseStudentOut, status_code=201)
async def join_course(invite_token: str, data: CourseJoinRequest, db: AsyncSession = Depends(get_db)):
    course = (await db.execute(select(Course).where(Course.invite_token == invite_token))).scalar_one_or_none()
    if not course or course.status == CourseStatus.archived:
        raise HTTPException(status_code=404, detail="Course not found")
    
    email_lower = str(data.email).strip().lower()
    password = data.password.strip()
    
    # Unified Account Logic:
    # Check if this user already has ANY registration with this email
    existing_reg = (await db.execute(
        select(CourseStudent).where(CourseStudent.email == email_lower)
    )).scalars().first()

    # Check if they are already registered for THIS specific course
    this_course_reg = (await db.execute(
        select(CourseStudent).where(
            CourseStudent.course_id == course.id, 
            CourseStudent.email == email_lower
        )
    )).scalar_one_or_none()

    if this_course_reg:
        if this_course_reg.hashed_password and not verify_password(password, this_course_reg.hashed_password):
            raise HTTPException(status_code=401, detail="Invalid email or password")
        if not this_course_reg.hashed_password:
            this_course_reg.hashed_password = hash_password(password)
            await db.commit()
            await db.refresh(this_course_reg)
        return await _student_out(this_course_reg, db)

    # Security: If user exists, they MUST provide correct password to add a new registration
    hashed_pass = None
    if existing_reg:
        if existing_reg.hashed_password:
            if not verify_password(password, existing_reg.hashed_password):
                raise HTTPException(status_code=401, detail="Пользователь с таким email уже существует. Пожалуйста, войдите в систему.")
            hashed_pass = existing_reg.hashed_password
        else:
            # Legacy account without password
            hashed_pass = hash_password(password)
    else:
        # New account
        hashed_pass = hash_password(password)

    student = CourseStudent(
        course_id=course.id,
        token=generate_viewer_token(),
        name=data.name.strip(),
        phone=data.phone.strip(),
        email=email_lower,
        telegram=data.telegram,
        hashed_password=hashed_pass,
        last_seen_at=datetime.utcnow(),
    )
    db.add(student)
    await db.commit()
    await db.refresh(student)
    return await _student_out(student, db)


@router.post("/course/login", response_model=list[CourseStudentOut])
async def student_login(data: dict, db: AsyncSession = Depends(get_db)):
    email = data.get("email", "").strip().lower()
    password = data.get("password", "")
    
    # Find registrations for this email
    result = await db.execute(select(CourseStudent).where(CourseStudent.email == email))
    regs = result.scalars().all()
    
    if not regs:
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
    
    # Check password (take the first one that has it)
    valid = False
    for r in regs:
        if r.hashed_password and verify_password(password, r.hashed_password):
            valid = True
            break
    
    if not valid:
        raise HTTPException(status_code=401, detail="Неверный email или пароль")
        
    return [await _student_out(r, db) for r in regs]


@router.put("/student/me")
async def update_student_profile(data: dict, email: str = Query(...), db: AsyncSession = Depends(get_db)):
    email_lower = email.strip().lower()
    # Update all registrations for this email
    result = await db.execute(select(CourseStudent).where(CourseStudent.email == email_lower))
    regs = result.scalars().all()
    if not regs:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    
    for reg in regs:
        if "name" in data: reg.name = data["name"]
        if "phone" in data: reg.phone = data["phone"]
        if "telegram" in data: reg.telegram = data["telegram"]
    
    await db.commit()
    return {"ok": True}


@router.put("/student/password")
async def update_student_password(data: dict, email: str = Query(...), db: AsyncSession = Depends(get_db)):
    email_lower = email.strip().lower()
    old_password = data.get("old_password")
    new_password = data.get("new_password")
    
    if not old_password or not new_password:
        raise HTTPException(status_code=400, detail="Необходимы старый и новый пароли")
        
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Новый пароль должен быть не менее 6 символов")

    result = await db.execute(select(CourseStudent).where(CourseStudent.email == email_lower))
    regs = result.scalars().all()
    if not regs:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
        
    # Verify old password against the first record that has one
    valid = False
    for r in regs:
        if r.hashed_password and verify_password(old_password, r.hashed_password):
            valid = True
            break
            
    if not valid:
        raise HTTPException(status_code=400, detail="Неверный текущий пароль")
        
    new_hashed = hash_password(new_password)
    for reg in regs:
        reg.hashed_password = new_hashed
        
    await db.commit()
    return {"ok": True}


@router.get("/student/dashboard", response_model=dict)
async def student_dashboard(tokens: list[str] = Query(default=[]), db: AsyncSession = Depends(get_db)):
    clean_tokens = [token.strip() for token in tokens if token.strip()]
    if not clean_tokens:
        raise HTTPException(status_code=401, detail="Student session is required")

    res = await db.execute(
        select(CourseStudent, Course)
        .join(Course, CourseStudent.course_id == Course.id)
        .where(CourseStudent.token.in_(clean_tokens))
    )
    rows = res.all()
    
    courses_data = []
    total_completed = 0
    total_score = 0
    quiz_count = 0
    
    for student, course in rows:
        student_data = await _student_out(student, db)
        
        # Calculate avg score for this course
        attempts_res = await db.execute(
            select(func.avg(QuizAttempt.score))
            .where(QuizAttempt.student_id == student.id, QuizAttempt.passed.is_(True))
        )
        avg_score = attempts_res.scalar() or 0
        
        courses_data.append({
            "course": await _course_out(course, db),
            "registration": student_data,
            "avg_quiz_score": round(float(avg_score), 1)
        })
        
        total_completed += student_data.completed_lessons
        if avg_score > 0:
            total_score += avg_score
            quiz_count += 1

    return {
        "courses": courses_data,
        "stats": {
            "total_courses": len(rows),
            "total_completed_lessons": total_completed,
            "avg_global_score": round(float(total_score / quiz_count), 1) if quiz_count else 0
        }
    }


@router.get("/course/{slug}/learn", response_model=CourseLearnOut)
async def learn_course(slug: str, token: str, db: AsyncSession = Depends(get_db)):
    course = (await db.execute(select(Course).where(Course.slug == slug))).scalar_one_or_none()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    student = (await db.execute(select(CourseStudent).where(CourseStudent.course_id == course.id, CourseStudent.token == token))).scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=403, detail="Invalid token")
        
    # Enforce access control
    if student.status != CourseStudentStatus.approved:
        raise HTTPException(status_code=403, detail="Доступ к курсу еще не подтвержден администратором.")
    
    student.last_seen_at = datetime.utcnow()
    
    modules = (await db.execute(
        select(CourseModule)
        .where(CourseModule.course_id == course.id, CourseModule.is_published.is_(True))
        .options(_quiz_load_options())
        .order_by(CourseModule.position, CourseModule.id)
    )).scalars().all()
    
    lessons = (
        await db.execute(
            select(CourseLesson)
            .join(CourseModule, CourseLesson.module_id == CourseModule.id)
            .where(CourseModule.course_id == course.id)
            .where(CourseModule.is_published.is_(True))
            .where(CourseLesson.is_published.is_(True))
            .order_by(CourseModule.position, CourseLesson.position, CourseLesson.id)
        )
    ).scalars().all()
    
    completed = (await db.execute(select(CourseLessonProgress.lesson_id).where(CourseLessonProgress.student_id == student.id))).scalars().all()
    
    # Quizzes passed
    passed_quizzes = (
        await db.execute(select(QuizAttempt.quiz_id).where(QuizAttempt.student_id == student.id, QuizAttempt.passed.is_(True)))
    ).scalars().all()
    quiz_ids = [module.quiz.id for module in modules if module.quiz]
    avg_quiz_score = None
    if quiz_ids:
        avg_quiz_score = (
            await db.execute(
                select(func.avg(QuizAttempt.score)).where(
                    QuizAttempt.student_id == student.id,
                    QuizAttempt.quiz_id.in_(quiz_ids),
                    QuizAttempt.passed.is_(True),
                )
            )
        ).scalar()
    
    await db.commit()
    return CourseLearnOut(
        course=await _course_out(course, db),
        student=await _student_out(student, db),
        modules=modules,
        lessons=lessons,
        completed_lesson_ids=list(completed),
        passed_quiz_ids=list(passed_quizzes),
        quiz_count=len(quiz_ids),
        avg_quiz_score=round(float(avg_quiz_score), 1) if avg_quiz_score is not None else None,
    )


@router.post("/course/lessons/{lesson_id}/complete")
async def complete_lesson(lesson_id: int, token: str, db: AsyncSession = Depends(get_db)):
    lesson = (await db.execute(select(CourseLesson).where(CourseLesson.id == lesson_id))).scalar_one_or_none()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    module = (await db.execute(select(CourseModule).where(CourseModule.id == lesson.module_id))).scalar_one()
    course = (await db.execute(select(Course).where(Course.id == module.course_id))).scalar_one()
    student = (await db.execute(select(CourseStudent).where(CourseStudent.course_id == module.course_id, CourseStudent.token == token))).scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=403, detail="Invalid token")
    await _ensure_student_can_learn(student)
    await _ensure_lesson_available(course, student, lesson, db)
    
    exists = (
        await db.execute(select(CourseLessonProgress).where(CourseLessonProgress.student_id == student.id, CourseLessonProgress.lesson_id == lesson_id))
    ).scalar_one_or_none()
    
    if not exists:
        db.add(CourseLessonProgress(student_id=student.id, lesson_id=lesson_id))
    
    student.last_seen_at = datetime.utcnow()
    await db.commit()
    return {"ok": True}


@router.post("/course/quizzes/{quiz_id}/attempt", response_model=QuizAttemptOut)
async def submit_quiz(quiz_id: int, token: str, data: QuizAttemptCreate, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Quiz).where(Quiz.id == quiz_id))
    quiz = result.scalar_one_or_none()
    if not quiz:
        raise HTTPException(status_code=404, detail="Quiz not found")
    
    module = (await db.execute(
        select(CourseModule).where(CourseModule.id == quiz.module_id)
    )).scalar_one_or_none()
    if not module or not module.is_published:
        raise HTTPException(status_code=404, detail="Module not found")
    course = (await db.execute(select(Course).where(Course.id == module.course_id))).scalar_one_or_none()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")

    student = (await db.execute(
        select(CourseStudent).where(
            CourseStudent.course_id == module.course_id,
            CourseStudent.token == token,
        )
    )).scalar_one_or_none()

    if not student:
        raise HTTPException(status_code=403, detail="Invalid token")
    await _ensure_student_can_learn(student)
    await _ensure_quiz_available(course, student, quiz, db)

    # Load questions and options
    questions = (await db.execute(select(QuizQuestion).where(QuizQuestion.quiz_id == quiz.id))).scalars().all()
    
    total_questions = len(questions)
    correct_count = 0
    
    for q in questions:
        # Load correct options for this question
        correct_opts = (await db.execute(
            select(QuizOption.id).where(QuizOption.question_id == q.id, QuizOption.is_correct.is_(True))
        )).scalars().all()
        
        student_answers = data.answers.get(str(q.id), [])
        
        if set(student_answers) == set(correct_opts):
            correct_count += 1
            
    score = round((correct_count / total_questions) * 100) if total_questions else 100
    passed = score >= quiz.passing_score
    
    attempt = QuizAttempt(
        student_id=student.id,
        quiz_id=quiz.id,
        score=score,
        passed=passed,
        answers_json=data.answers
    )
    db.add(attempt)
    await db.commit()
    await db.refresh(attempt)
    return attempt


@router.patch("/courses/students/{student_id}", response_model=CourseStudentOut)
async def update_student_status(
    student_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    student = (await db.execute(select(CourseStudent).where(CourseStudent.id == student_id))).scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    await get_course_for_user(student.course_id, user, db, "course.students")
    if "status" in data:
        try:
            student.status = CourseStudentStatus(data["status"])
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid student status") from None
    await db.commit()
    await db.refresh(student)
    return await _student_out(student, db)
