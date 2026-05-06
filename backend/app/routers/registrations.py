from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.models.user import User
from app.models.webinar import Webinar
from app.models.registration import Registration
from app.schemas.registration import InviteRegistrationCreate, RegistrationOut
from app.schemas.webinar import WebinarOut
from app.core.security import generate_viewer_token
from app.core.dependencies import get_current_user
from app.services.access import get_webinar_for_user

router = APIRouter(tags=["registrations"])


@router.get("/invite/{invite_token}", response_model=WebinarOut)
async def invite_info(
    invite_token: str,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Webinar).where(Webinar.invite_token == invite_token))
    webinar = result.scalar_one_or_none()
    if not webinar:
        raise HTTPException(status_code=404, detail="Invite link not found")
    return webinar


@router.post("/invite/{invite_token}/join", response_model=RegistrationOut, status_code=201)
async def join_by_invite(
    invite_token: str,
    data: InviteRegistrationCreate,
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(select(Webinar).where(Webinar.invite_token == invite_token))
    webinar = result.scalar_one_or_none()
    if not webinar:
        raise HTTPException(status_code=404, detail="Invite link not found")

    token = generate_viewer_token()
    reg = Registration(
        webinar_id=webinar.id,
        token=token,
        **data.model_dump(),
    )
    db.add(reg)
    await db.commit()
    await db.refresh(reg)
    return reg


@router.post("/webinars/{slug}/register")
async def register_by_slug_disabled(slug: str):
    raise HTTPException(status_code=404, detail="Use invite link")


@router.get("/admin/webinars/{webinar_id}/registrations", response_model=list[RegistrationOut])
async def list_registrations(
    webinar_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_webinar_for_user(webinar_id, current_user, db, permission="webinar.export")
    result = await db.execute(
        select(Registration).where(Registration.webinar_id == webinar_id).order_by(Registration.created_at.desc())
    )
    return result.scalars().all()


@router.get("/admin/webinars/{webinar_id}/registrations/export")
async def export_registrations(
    webinar_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_webinar_for_user(webinar_id, current_user, db, permission="webinar.export")
    result = await db.execute(
        select(Registration).where(Registration.webinar_id == webinar_id).order_by(Registration.created_at)
    )
    regs = result.scalars().all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "name", "phone", "email", "telegram", "token",
                     "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content", "created_at"])
    for r in regs:
        writer.writerow([
            r.id, r.name, r.phone, r.email, r.telegram, r.token,
            r.utm_source, r.utm_medium, r.utm_campaign, r.utm_term, r.utm_content,
            r.created_at.isoformat() if r.created_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=registrations_{webinar_id}.csv"},
    )


@router.get("/admin/webinars/{webinar_id}/registrations/export.xlsx")
async def export_registrations_xlsx(
    webinar_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await get_webinar_for_user(webinar_id, current_user, db, permission="webinar.export")

    result = await db.execute(
        select(Registration).where(Registration.webinar_id == webinar_id).order_by(Registration.created_at)
    )
    regs = result.scalars().all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Участники"

    headers = ["ФИО", "Телефон", "Email", "Telegram", "Дата входа"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="EEF2FF")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for reg in regs:
        sheet.append([
            reg.name,
            reg.phone,
            reg.email,
            reg.telegram,
            reg.created_at.strftime("%Y-%m-%d %H:%M:%S") if reg.created_at else "",
        ])

    widths = [32, 20, 32, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"participants_{webinar_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
